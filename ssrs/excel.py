from openpyxl import load_workbook
from ssrs.models import ParamReport


def read_xlsx_file(file_name: str, sheet_name: str = None) -> list[ParamReport]:
    """Чтение Excel файла с параметрами отчётов"""
    wb = load_workbook(file_name, data_only=True, read_only=True)
    sheet = wb.active if not sheet_name else wb[sheet_name]
    count_row, count_col = sheet.max_row, sheet.max_column

    columns_load = {
        'name': 1,
        'folder': 2,
        'report_name': 3,
        'params': 4,
        'to_path': 5,
        'to_email': 6,
        'blocked': 7,
        'format': 8,
    }

    result_lst = [
        ParamReport(
            name=str(sheet.cell(row=i, column=columns_load['name']).value),
            ssrs_folder=str(sheet.cell(row=i, column=columns_load['folder']).value),
            ssrs_report_name=str(sheet.cell(row=i, column=columns_load['report_name']).value),
            params=str(sheet.cell(row=i, column=columns_load['params']).value),
            to_path=str(sheet.cell(row=i, column=columns_load['to_path']).value),
            to_email=sheet.cell(row=i, column=columns_load['to_email']).value,
            blocked=sheet.cell(row=i, column=columns_load['blocked']).value,
            format=str(sheet.cell(row=i, column=columns_load['format']).value),
        )
        for i in range(2, count_row + 1)
        if sheet.cell(row=i, column=columns_load['blocked']).value is None
    ]
    wb.close()
    return result_lst
